import json
import sys
from pathlib import Path
from typing import List, Dict, Any

try:
    import openpyxl
    from openpyxl import Workbook
except ImportError:
    openpyxl = None

from src.services.page_emails import scrape_event_emails

OUTPUT_XLSX = Path("data/emails.xlsx")

COLUMNS = ["url", "organiser", "email", "source_json"]


def load_existing_urls(xlsx_path: Path) -> set:
    if not xlsx_path.exists() or openpyxl is None:
        return set()
    try:
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        existing = set()
        # Assume first row is header
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            url = (row[0] or "").strip()
            if url:
                existing.add(url)
        return existing
    except Exception:
        return set()


def ensure_workbook(xlsx_path: Path) -> "Workbook":
    if openpyxl is None:
        raise RuntimeError("openpyxl is required to write Excel. Install with: pip install openpyxl")
    if xlsx_path.exists():
        try:
            return openpyxl.load_workbook(xlsx_path)
        except Exception:
            pass
    wb = Workbook()
    ws = wb.active
    ws.title = "emails"
    ws.append(COLUMNS)
    return wb


def iter_events_in_folder(folder: Path) -> List[Dict[str, Any]]:
    """Recursively find all JSON files and extract events from them."""
    events = []
    # Search recursively for JSON files (events are in subdirs like relevant/, non-relevant/)
    for json_file in folder.rglob("*.json"):
        try:
            with open(json_file, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, list):
                for ev in data:
                    if isinstance(ev, dict):
                        ev_copy = dict(ev)
                        ev_copy["__source_json"] = str(json_file)
                        events.append(ev_copy)
            elif isinstance(data, dict):
                # Single event object
                ev_copy = dict(data)
                ev_copy["__source_json"] = str(json_file)
                events.append(ev_copy)
        except Exception:
            continue
    return events


def harvest(folder_path: str, use_playwright: bool = True) -> None:
    folder = Path(folder_path)
    if not folder.exists() or not folder.is_dir():
        print(f"Invalid folder: {folder}")
        return

    if openpyxl is None:
        print("openpyxl not installed; skipping Excel write. Install with: pip install openpyxl")
        return

    existing_urls = load_existing_urls(OUTPUT_XLSX)
    wb = ensure_workbook(OUTPUT_XLSX)
    ws = wb.active

    events = iter_events_in_folder(folder)
    added = 0
    skipped = 0

    total = len(events)
    for i, ev in enumerate(events):
        url = (ev.get("url") or ev.get("event_url") or "").strip()
        organiser = (ev.get("organiser") or ev.get("organizer") or ev.get("organiser_name") or "").strip()
        src = ev.get("__source_json", "")
        if not url:
            continue
        if url in existing_urls:
            skipped += 1
            continue

        # Show progress
        print(f"  [{i+1}/{total}] Scraping: {url[:60]}...")
        
        emails = scrape_event_emails(url, use_js=use_playwright)
        # Join all emails with comma (instead of just taking first one)
        email_str = ", ".join(emails) if emails else ""
        
        if emails:
            print(f"    ✓ Found {len(emails)} email(s): {email_str}")

        ws.append([url, organiser, email_str, src])
        existing_urls.add(url)
        added += 1

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(f"Done. Added {added} rows. Skipped {skipped} (already present). Saved to {OUTPUT_XLSX}.")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python -m src.utils.harvest_emails <events_output_folder> [--no-playwright]")
        sys.exit(1)
    folder = sys.argv[1]
    use_pw = True
    if len(sys.argv) > 2 and sys.argv[2] == "--no-playwright":
        use_pw = False
    harvest(folder, use_playwright=use_pw)
